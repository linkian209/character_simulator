'''
This program will simulate leveling a DnD character, showing their ending HP, and stats.
'''
import argparse
import csv
import json
import re
from openpyxl import load_workbook
from pandas import DataFrame


def import_race_data(file_path):
    '''
    This method imports data from the inputed CSV and returns a dictionary containing
    all of the data formated by race and subrace

    Arguments:
        :param import_data: (str) The filepath to the data

    Returns:
        dict: The dictionary of all of the data
    '''
    retval = {}

    # Open csv file and read in all data
    with open(file_path) as csv_file:
        reader = csv.DictReader(csv_file)
        for row in reader:
            race = row['Race']
            subrace = row['Subrace']

            if(subrace):
                if(race in retval):
                    if('Subraces' not in retval[race]):
                        retval[race]['Subraces'] = {}
                    retval[race]['Subraces'][subrace] = row
                else:
                    retval[race] = {'Subraces':{}}
                    retval[race]['Subraces'][subrace] = row
            else:
                retval[race] = row

    return retval

def update_mode(args):
    '''
    This method is the main method for running this program in Update mode.

    Update mode takes in a specifically formated XLSX file and outputs a JSON
    file containing all of the data for races and subraces needed by the
    program in run mode

    Arguments:
        :param args: (dict) A dictionary containing the needed arguments

    Returns:
        bool: Whether or not the update completed successfully or not
    '''
    # Lets first open the workbook
    try:
        workbook = load_workbook(args['xlsx_file'])
    except:
        return False

    # Now turn the Race sheet into a dataframe
    df = DataFrame()
    for name in workbook.sheetnames:
        if('Race' in name):
            df = DataFrame(workbook[name].values)

    # If we find nothing, return failure
    if(df.empty):
        return False

    # Lets remove the title row
    df.drop(0, axis=0, inplace=True)
    df.reset_index(inplace=True, drop=True)

    # Now lets get the headers, find the last column, and remove this row
    end_col = (df.iloc[0, :].values == None).argmax()
    df.drop(df.iloc[:, end_col:], axis=1, inplace=True)
    df.columns = list(df.iloc[0, :])
    df.drop(0, axis=0, inplace=True)
    df.reset_index(inplace=True, drop=True)

    # Now lets resize this dataframe to only contain the information we want
    # We first scroll down the rows to find the first blank cell, that is the
    # end of the rows
    end_row = (df.iloc[:, 0].values == None).argmax()
    df.drop(df[end_row:].index, axis=0, inplace=True)

    # Now let's get the race names and source names
    hyperlink_re = re.compile(r'(?<=,")(.+)(?=")')
    df['Race'] = df['Race'].apply(
        lambda x: x if hyperlink_re.search(x) is None else hyperlink_re.search(x).group(1)
    )
    df['Source'] = df['Source'].apply(
        lambda x: x if hyperlink_re.search(x) is None else hyperlink_re.search(x).group(1)
    )

    # Now make sure the stat fields are correct integers

    # Loop through dataframe so we can assemble the json in the format we want
    data = {}
    asi_re = re.compile(r'ASI: ([+-]\d) \(x(\d)\)(?:\s{1}\((.+)\))?')
    for index, row in df.iterrows():
        # First lets index this record into the correct spot in the array
        row = dict(row)
        race = row['Race']
        subrace = row['Subrace']

        if(subrace):
            if(race in data):
                if('Subraces' not in data[race]):
                    data[race]['Subraces'] = {}
                data[race]['Subraces'][subrace] = row
            else:
                data[race] = {'Subraces':{}}
                data[race]['Subraces'][subrace] = row
        else:
            data[race] = row

        # Now that we have added this row, check if there are any special ASI rules to note
        if(row['Additional'] is not None):
            matches = asi_re.search(row['Additional'])
            if(matches):
                # We found something
                asi = {'size': matches.group(1), 'number': matches.group(2)}

                # Check if we have restrictions
                if(matches.group(3)):
                    # We either can put the point into a number of options, or not
                    # into one stat
                    if('-' in matches.group(3)):
                        # We cannot use this stat
                        asi['not_allowed'] = matches.group(3).split('-')[1]

                    if('|' in matches.group(3)):
                        # We can only use one or the other
                        asi['allowed'] = [x.capitalize() for x in matches.group(3).split(' | ')]
                
                # Now add this to the row of data
                if(subrace):
                    data[race]['Subraces'][subrace]['ASI'] = asi
                else:
                    data[race]['ASI'] = asi

    # Done! Let's dump this file
    with open('race_data.json', 'w') as fp:
        json.dump(data, fp, indent=2)

    return True

def run_mode(args):
    '''
    This method is the main method for running this program in Run mode.

    This mode goes through the character simulation

    Arguments:
        :param args: (dict) A dictionary containing the needed arguments
    '''
    pass


if __name__ == "__main__":
    # Setup argument parsers and parse arguments
    main_parser = argparse.ArgumentParser(description='Character Simulator')
    subparsers = main_parser.add_subparsers(help='Mode Help')

    update_parser = subparsers.add_parser('update', help='Update Help')
    update_parser.add_argument('xlsx_file', type=str, help='Path to the .xlsx race file')

    run_parser = subparsers.add_parser('run', help='Run Help')

    args = vars(main_parser.parse_args())

    # If we are in update mode, update the json file
    if('xlsx_file' in args):
        update_mode(args)
    else:
        run_mode(args)